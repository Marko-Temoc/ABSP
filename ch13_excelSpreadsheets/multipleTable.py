#!/usr/bin/env python3

import openpyxl, sys
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

input = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
foont = Font(bold=True)

col_tracker1 = 2

#set up first row to multiply with later
for i in range(1, input + 1):
    #write i to B1, then C1, then D1, etc.
    sheet[f'{get_column_letter(col_tracker1)}1'] = i
    #apply bold font to row 1 values getting filled
    sheet[f'{get_column_letter(col_tracker1)}1'].font = foont
    #move onto next column for row 1
    col_tracker1 += 1
#fill col a with values to multiply with later
for j in range(1, input + 1):
    sheet[f'A{j + 1}'] = j
    x = sheet[f'A{j + 1}'].value
    sheet[f'A{j + 1}'].font = foont
    col_tracker2 = 2
    #fill out the row for every row
    for k in range(1,  input + 1):
        y = sheet[f'{get_column_letter(col_tracker2)}1'].value
        sheet[f'{get_column_letter(col_tracker2)}{j + 1}'] = (x * y)
        col_tracker2 += 1

wb.save('timesTable.xlsx')

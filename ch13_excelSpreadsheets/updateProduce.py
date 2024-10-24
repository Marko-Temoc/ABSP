#!/usr/bin/env python3
#updateProduce.py - Corrects costs in produce sales spreadsheet
import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
#the produce types and their updates prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}
#TODO loop through the rows and update the prices
for rowNum in range(2, sheet.max_row): #skip first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updateProduceSales.xlsx')
